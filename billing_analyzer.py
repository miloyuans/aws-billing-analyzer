#!/usr/bin/env python3
# File: aws-billing-analyzer/billing_analyzer.py
import os
import boto3
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ==================== 基础信息 ====================
sts = boto3.client('sts')
identity = sts.get_caller_identity()
ACCOUNT_ID = identity['Account']
try:
    ALIAS = boto3.client('iam').list_account_aliases()['AccountAliases'][0]
except:
    ALIAS = f"account-{ACCOUNT_ID}"

today = datetime.utcnow().date()
if today.day == 1:
    target_month = (today - timedelta(days=1)).strftime('%Y%m')
else:
    target_month = today.strftime('%Y%m')

FILE_PATH = f"/tmp/{ACCOUNT_ID}_{ALIAS}_{target_month}_Billing.xlsx"

# ==================== 数据获取 ====================
def get_cost_data():
    ce = boto3.client('ce', region_name='us-east-1')
    start_date = (today.replace(day=1) if today.day > 1 else (today - timedelta(days=1)).replace(day=1)).strftime('%Y-%m-%d')
    end_date = (today if today.day > 1 else today - timedelta(days=1)).strftime('%Y-%m-%d')

    response = ce.get_cost_and_usage(
        TimePeriod={'Start': start_date, 'End': end_date},
        Granularity='DAILY',
        Metrics=['UnblendedCost'],
        GroupBy=[
            {'Type': 'DIMENSION', 'Key': 'SERVICE'},
            {'Type': 'TAG', 'Key': 'Project'},
            {'Type': 'TAG', 'Key': 'Environment'}
        ]
    )

    rows = []
    for period in response['ResultsByTime']:
        date = period['TimePeriod']['Start']
        for group in period['Groups']:
            keys = group['Keys']
            service = keys[0]
            project = keys[1].split('$')[-1] if len(keys)>1 and '$' in keys[1] else 'Untagged'
            env = keys[2].split('$')[-1] if len(keys)>2 and '$' in keys[2] else 'Untagged'
            cost = float(group['Metrics']['UnblendedCost']['Amount'])
            rows.append({'Date': date, 'Service': service, 'Project': project, 'Environment': env, 'Cost_USD': round(cost, 4)})
    return pd.DataFrame(rows)

# ==================== 主函数 ====================
def main():
    df = get_cost_data()
    total_cost = df['Cost_USD'].sum()

    # 1. 加载或创建工作簿
    if os.path.exists(FILE_PATH):
        wb = load_workbook(FILE_PATH)
        if '明细_Daily' in wb.sheetnames:
            ws_daily = wb['明细_Daily']
            for row in ws_daily.iter_rows(min_row=2):
                for cell in row: cell.value = None
        else:
            ws_daily = wb.create_sheet('明细_Daily', 0)
    else:
        wb = load_workbook(filename=os.devnull) if False else load_workbook()
        ws_daily = wb.active
        ws_daily.title = '明细_Daily'

    # 写入表头
    headers = ['日期', '服务', '项目', '环境', '费用(USD)']
    ws_daily.append(headers)

    # 写入明细数据
    for _, row in df.iterrows():
        ws_daily.append([row['Date'], row['Service'], row['Project'], row['Environment'], row['Cost_USD']])

    # 2. 创建/刷新汇总页（含4张图表）
    if '汇总_Summary' in wb.sheetnames:
        wb.remove(wb['汇总_Summary'])
    ws = wb.create_sheet('汇总_Summary')

    # 标题
    ws['A1'] = f"AWS 月度账单报告 - {target_month[:4]}年{int(target_month[4:]):02d}月"
    ws['A1'].font = ws['A1'].font.copy(size=18, bold=True)
    ws['A2'] = f"账户ID：{ACCOUNT_ID} ({ALIAS})"
    ws['A3'] = f"统计周期：{df['Date'].min()} 至 {df['Date'].max()}"
    ws['A4'] = f"本月累计费用：${total_cost:,.4f}"
    ws['A4'].font = ws['A4'].font.copy(size=16, bold=True, color="FF0000")

    # 图表1：每日趋势折线图
    daily_sum = df.groupby('Date')['Cost_USD'].sum().reset_index()
    ws.append([]); ws.append(['每日费用趋势'])
    ws.append(['日期', '费用'])
    for _, r in daily_sum.iterrows():
        ws.append([r['Date'], r['Cost_USD']])

    chart1 = LineChart()
    chart1.title = "本月每日费用趋势"
    chart1.y_axis.title = "费用 (USD)"
    chart1.x_axis.title = "日期"
    data = Reference(ws, min_col=2, min_row=ws.max_row-len(daily_sum), max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=ws.max_row-len(daily_sum)+1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "A10")

    # 图表2：服务占比饼图
    service_top = df.groupby('Service')['Cost_USD'].sum().nlargest(8)
    start_row = ws.max_row + 5
    ws.cell(row=start_row, column=10, value="服务费用占比 Top8")
    for i, (svc, cost) in enumerate(service_top.items(), start_row+2):
        ws.cell(row=i, column=10, value=svc)
        ws.cell(row=i, column=11, value=cost)

    pie = PieChart()
    data_pie = Reference(ws, min_col=11, min_row=start_row+1, max_row=start_row+1+len(service_top))
    cats_pie = Reference(ws, min_col=10, min_row=start_row+2, max_row=start_row+1+len(service_top))
    pie.add_data(data_pie)
    pie.set_categories(cats_pie)
    pie.title = "服务费用占比"
    ws.add_chart(pie, "J10")

    # 图表3：项目柱状图
    project_top = df.groupby('Project')['Cost_USD'].sum().nlargest(10)
    row = start_row + 10
    ws.cell(row=row, column=1, value="项目费用排名 Top10")
    for i, (proj, cost) in enumerate(project_top.items(), row+2):
        ws.cell(row=i, column=1, value=proj)
        ws.cell(row=i, column=2, value=cost)

    bar = BarChart()
    data_bar = Reference(ws, min_col=2, min_row=row+1, max_row=row+1+len(project_top))
    cats_bar = Reference(ws, min_col=1, min_row=row+2, max_row=row+1+len(project_top))
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(cats_bar)
    bar.title = "项目费用排名"
    ws.add_chart(bar, f"A{row+3}")

    # 保存
    wb.save(FILE_PATH)
    print(FILE_PATH)

if __name__ == "__main__":
    main()